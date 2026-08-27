from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any


SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}


def parse_orders_spreadsheet(content: bytes, filename: str) -> list[dict[str, Any]]:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        rows = _parse_csv(content)
    elif extension == ".xlsx":
        rows = _parse_xlsx(content)
    else:
        raise ValueError("Unsupported spreadsheet format. Use .csv or .xlsx.")

    orders = [_normalize_order_row(row, index=index) for index, row in enumerate(rows, start=1)]
    if not orders:
        raise ValueError("Spreadsheet has no order rows.")
    return orders


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_key(value) for value in rows[0]]
    parsed: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        parsed.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    return parsed


def _normalize_order_row(row: dict[str, Any], *, index: int) -> dict[str, Any]:
    normalized = {_normalize_key(key): value for key, value in row.items()}
    order_id = _text(_first(normalized, "order_id", "id", "pedido_id", "pedido")) or f"ORDER-{index:03d}"
    delivery_lat = _float(_required(normalized, "delivery_lat", "dropoff_lat", "lat_delivery", "lat_entrega"))
    delivery_lng = _float(
        _required(normalized, "delivery_lng", "delivery_lon", "dropoff_lng", "dropoff_lon", "lng_delivery", "lon_delivery", "lng_entrega")
    )
    pickup_lat = _float(_first(normalized, "pickup_lat", "origin_lat", "lat_pickup", "lat_coleta", default=delivery_lat))
    pickup_lng = _float(
        _first(
            normalized,
            "pickup_lng",
            "pickup_lon",
            "origin_lng",
            "origin_lon",
            "lng_pickup",
            "lon_pickup",
            "lng_coleta",
            default=delivery_lng,
        )
    )
    return {
        "id": order_id,
        "pickup": {
            "address": _text(_first(normalized, "pickup_address", "origin_address", "endereco_coleta", default="Pickup")),
            "lat": pickup_lat,
            "lng": pickup_lng,
        },
        "delivery": {
            "address": _text(
                _first(normalized, "delivery_address", "dropoff_address", "address", "endereco_entrega", default=f"Delivery {index}")
            ),
            "lat": delivery_lat,
            "lng": delivery_lng,
        },
        "weight": _float(_first(normalized, "weight", "peso", "demand", "demanda", default=1)),
        "weight_unit": _text(_first(normalized, "weight_unit", "unidade_peso", default="kg")),
        "volume": _optional_float(_first(normalized, "volume", "vol", "cubagem", default=None)),
        "volume_unit": _optional_text(_first(normalized, "volume_unit", "unidade_volume", default=None)),
        "priority": int(_float(_first(normalized, "priority", "prioridade", default=1))),
        "description": _optional_text(_first(normalized, "description", "descricao", "notes", "observacoes", default=None)),
    }


def _normalize_key(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def _first(row: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip() != "":
            return value
    return default


def _required(row: dict[str, Any], *keys: str) -> Any:
    value = _first(row, *keys, default=None)
    if value is None:
        raise ValueError(f"Missing required column. Expected one of: {', '.join(keys)}")
    return value


def _text(value: Any) -> str:
    return str(value or "").strip()


def _optional_text(value: Any) -> str | None:
    text = _text(value)
    return text or None


def _float(value: Any) -> float:
    if isinstance(value, str):
        value = value.strip().replace(",", ".")
    return float(value)


def _optional_float(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    return _float(value)
